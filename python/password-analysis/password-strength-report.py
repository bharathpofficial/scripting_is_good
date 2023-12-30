try:
    from zxcvbn import zxcvbn
    import openpyxl as xl
    from pathlib import Path
    import sys, os, argparse
except ModuleNotFoundError as err:
    print(f"requried modules not found!!\npathlib\nzxcvbn\nopenpyxl\npathlib\nargparse")
# ============================================================================
class PasswordsAnalysis:
    def __init__(self) -> None:
        self.test_password = ''
        self.saveas = ''
        # creats instance of xlfile class, whenever, PasswordAnalysis() is created
        self.xlobj = xlfile()
        self.__headers = []
        self.__values = []
        self.__raw = dict()
        print(self.saveas)

    def __cleanify(self) -> None:
        '''
        Intended to be auxillary method, for the below method.
        '''
        __results = zxcvbn(self.test_password)
        self.__raw = {
            'Password' : __results.get('password'),
            # 'calc' : __results.get('calc_time'),
            'Score' : __results.get('score'),
            # 'suggestions' : __results.get('feedback').get('suggestions')[0],
            'Pattern' : __results.get('sequence')[0].get('pattern'),
            'online_slow' : __results.get('crack_times_display').get('online_throttling_100_per_hour'),
            'online_fast' : __results.get('crack_times_display').get('online_no_throttling_10_per_second'),
            'offline_slow' : __results.get('crack_times_display').get('offline_slow_hashing_1e4_per_second'),
            'offline_fast' : __results.get('crack_times_display').get('offline_fast_hashing_1e10_per_second'),
            'isleet' : __results.get('sequence')[0].get('l33t'),
            'warning' : __results.get('feedback').get('warning')
        }
        self.__headers = list(self.__raw.keys())
        self.__values = list(self.__raw.values())

    def __init_file(self) -> None:
        __filename = Path(self.xlobj.saveas)
        if not __filename.exists():
            print(f'filename: {self.xlobj.saveas}')
            if str(__filename.suffix).lower() == '':
                self.xlobj.saveas = self.xlobj.saveas + '.xlsx'
            self.xlobj.createworkbook()
            self.xlobj._upd_shName(self.xlobj.sheetName)
            self.xlobj.workbook.active.append(self.__headers)
            self.xlobj.saveFile(self.xlobj.saveas)

        elif __filename.exists():
            if str(__filename.suffix).lower() == '.xlsx':
                self.xlobj._getWorkbook(existfile=self.xlobj.saveas)
            else:
                print(f'Close file, then RUN. if you Opened {self.xlobj.saveas} ')
                sys.exit(1)

    def __generate_test_report(self) -> None:
        self.xlobj.workbook.active.append(self.__values)

    def report_password_strength(self) -> None:
        self.__cleanify()
        self.__init_file()
        self.__generate_test_report()
        self.xlobj.saveFile(self.xlobj.saveas)

# ============================================================================

# ============================================================================

class xlfile:
    def __init__(self) -> None:
        self.sheetName='Password_Report'
        self.workbook = None
        self.saveas=f'Passwords_Analysis.xlsx'

    def createworkbook(self):
        self.workbook = xl.Workbook()

    def _getWorkbook(self,existfile) -> None:
        self.workbook = xl.load_workbook(existfile)
        self.workbook.active.title = self.sheetName

    def _upd_shName(self,sheetName) -> None:
        self.workbook.active.title = sheetName

    def saveFile(self,saveas) -> None:
        self.workbook.save(self.saveas) # to save the output:must use!!

def checkfile() -> None:
    parser = argparse.ArgumentParser(description='Get Your passwords analysed and get exported as excelfile.')
    parser.add_argument('input_file', help='the input file')
    args = parser.parse_args()
    if not os.path.exists(args.input_file):
        print(f'Make sure that you have the file: {args.input_file} you claim to have. !!')
        sys.exit(1)
    elif os.path.getsize(args.input_file) == 0:
        print(f'empty file.')
        sys.exit(1)
    return args.input_file

def operate(file_path):
    try:
        with open(file_path,'r') as passfile:
            for passwd in passfile:
                analyse.test_password = passwd
                analyse.report_password_strength()
    except FileNotFoundError as err:
        print(f'check this: {str(err)}')
# ============================================================================
if __name__ == "__main__":
    analyse = PasswordsAnalysis()
    file_path = checkfile()
    print('Loading...')
    operate(file_path)
    